from fastapi import APIRouter, Depends, Query
from typing import Annotated
from sqlalchemy import select, func, extract
from sqlalchemy.ext.asyncio import AsyncSession
from app.dependencies import require_staff_or_admin
from app.database import get_db
from app.models.user import User
from app.models.request import Request
from app.models.approval import Approval

router = APIRouter(prefix="/reports", tags=["Reporting & Analytics"])


@router.get("/summary", summary="Counts by request status")
async def summary(
    staff: Annotated[User, Depends(require_staff_or_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    rows = await db.execute(
        select(Request.status, func.count().label("count"))
        .where(Request.status != "draft")
        .group_by(Request.status)
        .order_by(Request.status)
    )
    status_counts = {row.status: row.count for row in rows}
    total = sum(status_counts.values())

    return {"total": total, "status_counts": status_counts}


@router.get("/trends", summary="Monthly request trends")
async def trends(
    staff: Annotated[User, Depends(require_staff_or_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
    year: int = Query(2026, ge=2020, le=2100),
):
    rows = await db.execute(
        select(
            extract("month", Request.created_at).label("month"),
            func.count().label("count"),
        )
        .where(
            Request.status != "draft",
            extract("year", Request.created_at) == year,
        )
        .group_by("month")
        .order_by("month")
    )
    monthly = {int(row.month): row.count for row in rows}
    data = [monthly.get(m, 0) for m in range(1, 13)]

    return {"year": year, "monthly": data}


@router.get("/by-program", summary="Requests grouped by programme")
async def by_program(
    staff: Annotated[User, Depends(require_staff_or_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from app.models.student import Student
    rows = await db.execute(
        select(Student.program, func.count(Request.request_id).label("count"))
        .join(Request, Request.student_id == Student.student_id)
        .where(Request.status != "draft")
        .group_by(Student.program)
        .order_by(func.count(Request.request_id).desc())
    )
    return [{"program": r.program, "count": r.count} for r in rows]


@router.get("/approval-timeline", summary="Average approval time")
async def approval_timeline(
    staff: Annotated[User, Depends(require_staff_or_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    rows = await db.execute(
        select(
            Approval.approver_role,
            func.avg(
                func.extract("epoch", Approval.decided_at - Request.submitted_at)
            ).label("avg_seconds"),
            func.count().label("count"),
        )
        .join(Request, Request.request_id == Approval.request_id)
        .group_by(Approval.approver_role)
    )
    result = []
    for row in rows:
        role = row.approver_role
        avg_seconds = row.avg_seconds
        avg_hours = round(avg_seconds / 3600, 1) if avg_seconds else None
        result.append({
            "role": role,
            "average_hours": avg_hours,
            "total_decisions": row.count,
        })
    return result


@router.get("/export", summary="Export requests as CSV, XLSX, or PDF")
async def export(
    format: str = Query("csv", regex="^(csv|xlsx|pdf)$"),
    status: str = Query(None, description="Filter by status"),
    staff: User = Depends(require_staff_or_admin),
    db: AsyncSession = Depends(get_db),
):
    from fastapi.responses import StreamingResponse
    from app.models.student import Student
    from sqlalchemy.orm import selectinload
    import csv, io, datetime

    q = (
        select(Request)
        .options(selectinload(Request.student).selectinload(Student.user))
        .where(Request.status != "draft")
        .order_by(Request.created_at.desc())
    )
    if status:
        q = q.where(Request.status == status)

    rows = await db.execute(q)
    requests = rows.scalars().all()

    def row_data(r):
        student_name = r.student.user.name if r.student and r.student.user else ""
        student_num = r.student.student_number if r.student else ""
        program = r.student.program if r.student else ""
        return {
            "request_id": str(r.request_id),
            "student": student_name,
            "student_number": student_num,
            "program": program,
            "reason": r.reason,
            "status": r.status,
            "scope": r.scope,
            "academic_year": r.academic_year,
            "semester": r.semester,
            "submitted_at": r.submitted_at.strftime("%Y-%m-%d %H:%M") if r.submitted_at else "",
        }

    if format == "csv":
        def csv_stream():
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(["Request ID", "Student", "Student #", "Program", "Reason", "Status", "Scope", "Academic Year", "Semester", "Submitted At"])
            for r in requests:
                d = row_data(r)
                w.writerow([d["request_id"], d["student"], d["student_number"], d["program"], d["reason"], d["status"], d["scope"], d["academic_year"], d["semester"], d["submitted_at"]])
            yield buf.getvalue()
        return StreamingResponse(
            csv_stream(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=requests.csv"},
        )

    elif format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import PieChart, BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.utils import get_column_letter
        from collections import Counter

        wb = Workbook()
        ws = wb.active
        ws.title = "Requests"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = ["Request ID", "Student", "Student #", "Program", "Reason", "Status", "Scope", "Academic Year", "Semester", "Submitted At"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        keys = ["request_id", "student", "student_number", "program", "reason", "status", "scope", "academic_year", "semester", "submitted_at"]
        for i, r in enumerate(requests, 2):
            d = row_data(r)
            for col, key in enumerate(keys, 1):
                cell = ws.cell(row=i, column=col, value=d.get(key, ""))
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ── Chart sheet ──────────────────────────────────────────────
        cs = wb.create_sheet("Charts")

        # Pie chart data
        status_counts = Counter(r.status for r in requests)
        status_order = ["approved", "rejected", "pending_hod", "pending_hod_exams", "pending_manager", "ineligible", "queried"]
        cs["A1"] = "Status"
        cs["B1"] = "Count"
        row_idx = 2
        for s in status_order:
            if status_counts.get(s, 0) > 0:
                cs.cell(row=row_idx, column=1, value=s.replace("_", " ").title())
                cs.cell(row=row_idx, column=2, value=status_counts[s])
                row_idx += 1
        pie = PieChart()
        pie.title = "Requests by Status"
        pie.style = 10
        pie.width = 16
        pie.height = 10
        data_ref = Reference(cs, min_col=1, max_col=2, min_row=1, max_row=row_idx - 1)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(Reference(cs, min_col=1, min_row=2, max_row=row_idx - 1))
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        cs.add_chart(pie, "D1")

        # Bar chart data
        from collections import defaultdict
        monthly = defaultdict(int)
        for r in requests:
            if r.created_at:
                key = r.created_at.strftime("%b-%Y")
                monthly[key] += 1
        col_offset = row_idx + 3
        cs.cell(row=col_offset, column=1, value="Month")
        cs.cell(row=col_offset, column=2, value="Count")
        for i, (month, count) in enumerate(sorted(monthly.items(), key=lambda x: x[0]), col_offset + 1):
            cs.cell(row=i, column=1, value=month)
            cs.cell(row=i, column=2, value=count)
        bar = BarChart()
        bar.type = "col"
        bar.title = "Monthly Request Trends"
        bar.style = 10
        bar.width = 16
        bar.height = 10
        bar.y_axis.title = "Count"
        bar_data = Reference(cs, min_col=1, max_col=2, min_row=col_offset, max_row=col_offset + len(monthly))
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(Reference(cs, min_col=1, min_row=col_offset + 1, max_row=col_offset + len(monthly)))
        bar.series[0].graphicalProperties.solidFill = "4472C4"
        cs.add_chart(bar, f"D{col_offset}")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=requests.xlsx"},
        )

    else:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
        from collections import Counter

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Requests Report",
                                topMargin=20, bottomMargin=20, leftMargin=20, rightMargin=20)
        styles = getSampleStyleSheet()
        elements = []
        page_width = landscape(A4)[0] - 40

        elements.append(Paragraph("Academic Postponement Requests Report", styles["Title"]))
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
        elements.append(Spacer(1, 16))

        # ── Compute chart data ──────────────────────────────────────
        status_counts = Counter(r.status for r in requests)
        status_order = ["approved", "rejected", "pending_hod", "pending_hod_exams", "pending_manager", "ineligible", "queried"]
        chart_labels = [s.replace("_", " ").title() for s in status_order if status_counts.get(s, 0) > 0]
        chart_values = [status_counts.get(s, 0) for s in status_order if status_counts.get(s, 0) > 0]

        # Monthly trends (last 12 months from latest request)
        from collections import defaultdict
        monthly = defaultdict(int)
        for r in requests:
            if r.created_at:
                key = r.created_at.strftime("%b %Y")
                monthly[key] += 1
        month_labels = list(monthly.keys())
        month_values = list(monthly.values())

        # ── Generate charts via matplotlib ───────────────────────────
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11, 4))

        # Pie chart
        pie_colors = ["#4CAF50", "#F44336", "#FF9800", "#FF9800", "#FF9800", "#9E9E9E", "#2196F3"]
        wedges, texts, autotexts = ax1.pie(
            chart_values, labels=None, autopct="%1.0f%%",
            colors=pie_colors[:len(chart_values)],
            startangle=90, textprops={"fontsize": 9}
        )
        ax1.set_title("Requests by Status", fontsize=12, fontweight="bold", pad=12)
        ax1.axis("equal")
        legend_labels = [f"{l} ({v})" for l, v in zip(chart_labels, chart_values)]
        ax1.legend(wedges, legend_labels, loc="lower center", bbox_to_anchor=(0.5, -0.22),
                   ncol=min(len(chart_labels), 4), fontsize=8, frameon=False)

        # Bar chart
        if month_labels:
            x_pos = range(len(month_labels))
            bars = ax2.bar(x_pos, month_values, color="#4472C4", width=0.6, edgecolor="white")
            ax2.set_title("Monthly Request Trends", fontsize=12, fontweight="bold", pad=12)
            ax2.set_xticks(x_pos)
            ax2.set_xticklabels(month_labels, fontsize=7, rotation=25, ha="right")
            ax2.set_ylabel("Count", fontsize=9)
            for bar, val in zip(bars, month_values):
                ax2.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
                         str(val), ha="center", va="bottom", fontsize=8)
        else:
            ax2.text(0.5, 0.5, "No data", ha="center", va="center", fontsize=12, transform=ax2.transAxes)
            ax2.set_title("Monthly Request Trends", fontsize=12, fontweight="bold", pad=12)

        plt.tight_layout(pad=2)
        chart_buf = io.BytesIO()
        fig.savefig(chart_buf, format="png", dpi=150, bbox_inches="tight")
        plt.close(fig)
        chart_buf.seek(0)
        elements.append(RLImage(chart_buf, width=page_width, height=page_width * 0.36))
        elements.append(Spacer(1, 16))

        # ── Summary KPIs ─────────────────────────────────────────────
        total = len(requests)
        kpi_data = [
            ["Total", "Approved", "Rejected", "Pending", "Ineligible"],
            [
                str(total),
                str(status_counts.get("approved", 0)),
                str(status_counts.get("rejected", 0)),
                str(status_counts.get("pending_hod", 0) + status_counts.get("pending_hod_exams", 0) + status_counts.get("pending_manager", 0)),
                str(status_counts.get("ineligible", 0)),
            ],
        ]
        kpi_table = Table(kpi_data, colWidths=[page_width / 5] * 5)
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.Color(0.27, 0.45, 0.77)),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("BACKGROUND", (0, 1), (-1, 1), rl_colors.Color(0.92, 0.94, 0.98)),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 16))

        # ── Detail table ─────────────────────────────────────────────
        elements.append(Paragraph("Request Details", styles["Heading2"]))
        elements.append(Spacer(1, 8))

        table_data = [["ID", "Student", "Program", "Reason", "Status", "Submitted"]]
        for r in requests:
            d = row_data(r)
            table_data.append([
                str(r.request_id)[:8],
                d["student"],
                d["program"],
                (d["reason"][:50] + "...") if len(d["reason"]) > 50 else d["reason"],
                d["status"],
                d["submitted_at"],
            ])

        status_colors = {
            "approved": rl_colors.Color(0.16, 0.77, 0.10, 0.15),
            "rejected": rl_colors.Color(1, 0.30, 0.30, 0.15),
            "pending_hod": rl_colors.Color(1, 0.68, 0.08, 0.15),
            "pending_hod_exams": rl_colors.Color(1, 0.68, 0.08, 0.15),
            "pending_manager": rl_colors.Color(1, 0.68, 0.08, 0.15),
        }

        col_widths = [page_width * 0.12, page_width * 0.16, page_width * 0.16, page_width * 0.28, page_width * 0.14, page_width * 0.14]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.Color(0.27, 0.45, 0.77)),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for idx, r in enumerate(requests):
            row_idx = idx + 1
            bg = status_colors.get(r.status)
            if bg:
                style_cmds.append(("BACKGROUND", (4, row_idx), (4, row_idx), bg))

        t.setStyle(TableStyle(style_cmds))
        elements.append(t)

        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=requests.pdf"},
        )
